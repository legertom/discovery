"""
Build AI_Enablement_Intake.xlsx — comprehensive workflow friction intake workbook.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Design tokens
# ---------------------------------------------------------------------------
NAVY        = "1F4E79"
NAVY_FONT   = "FFFFFF"
MID_BLUE    = "2E75B6"
ROW_ALT     = "EBF3FA"
ROW_WHITE   = "FFFFFF"
BORDER_COLOR= "BFBFBF"
GREEN_FILL  = "C6EFCE"; GREEN_FONT  = "375623"
YELLOW_FILL = "FFEB9C"; YELLOW_FONT = "9C6500"
RED_FILL    = "FFC7CE"; RED_FONT    = "9C0006"
PURPLE_FILL = "E2CFED"; PURPLE_FONT = "7030A0"
BLUE_FILL   = "BDD7EE"; BLUE_FONT   = "1F4E79"
GRAY_FILL   = "D9D9D9"; GRAY_FONT   = "595959"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(cell, text, fill_hex="1F4E79", font_hex="FFFFFF", bold=True, size=10):
    cell.value = text
    cell.font = Font(name='Calibri', bold=bold, color=font_hex, size=size)
    cell.fill = make_fill(fill_hex)
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell.border = make_border()

def style_data(cell, fill_hex="FFFFFF"):
    cell.font = Font(name='Calibri', size=10)
    cell.fill = make_fill(fill_hex)
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    cell.border = make_border()

def add_note(cell, text):
    c = Comment(text, "System")
    c.width = 300
    c.height = 100
    cell.comment = c

def apply_alt_rows(ws, start_row, end_row, num_cols):
    for r in range(start_row, end_row + 1):
        fill_hex = ROW_ALT if r % 2 == 1 else ROW_WHITE
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            rgb = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else ""
            if rgb in ("00000000", "FFFFFFFF", "FFFFFF", ""):
                cell.fill = make_fill(fill_hex)
            cell.border = make_border()
            cell.alignment = Alignment(wrap_text=True, vertical='top')

def set_col_widths(ws, widths_dict):
    """widths_dict: {col_letter_or_num: width}"""
    for col, w in widths_dict.items():
        if isinstance(col, int):
            col = get_column_letter(col)
        ws.column_dimensions[col].width = w

def make_dv(formula1, type_="list", showDropDown=False):
    dv = DataValidation(type=type_, formula1=formula1, allow_blank=True)
    dv.showDropDown = showDropDown
    return dv

def apply_range_dv(ws, dv, col_letter, start_row, end_row):
    dv.sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.add_data_validation(dv)

# ---------------------------------------------------------------------------
# Workbook setup — sheets in required order
# ---------------------------------------------------------------------------
wb = Workbook()
# Remove default sheet
wb.remove(wb.active)

# Sheet order
ws_intake    = wb.create_sheet("Intake Responses")
ws_scoring   = wb.create_sheet("Opportunity Scoring")
ws_discovery = wb.create_sheet("Discovery Sessions")
ws_steplog   = wb.create_sheet("Discovery Step Log")
ws_briefs    = wb.create_sheet("Opportunity Briefs")
ws_dashboard = wb.create_sheet("Dashboard")
ws_lists     = wb.create_sheet("Lists & Settings")
ws_instr     = wb.create_sheet("Instructions")

# ---------------------------------------------------------------------------
# SHEET 7: Lists & Settings  (build first — other sheets reference it)
# ---------------------------------------------------------------------------
ws_lists.sheet_properties.tabColor = "595959"

lists_headers = [
    "Team / Function", "Frequency", "Friction Type",
    "Likely Solution Type", "Opportunity Status",
    "Risk Level", "Feasibility", "Priority Category"
]
for i, h in enumerate(lists_headers, 1):
    style_header(ws_lists.cell(1, i), h)

col_A = ["Customer Support","Customer Success","Product","Engineering","Sales",
         "Marketing","Operations","People","Finance","Legal","Security","IT",
         "Leadership","Other"]
col_B = ["Multiple times per day","Daily","Weekly","Monthly","Quarterly","Ad hoc","Unknown"]
col_C = ["Repetitive copy/paste","Multi-system data gathering","Manual reporting",
         "Writing or rewriting","Summarizing or synthesizing information",
         "Searching for information","Data cleanup or normalization",
         "Manual QA/checking","Slow handoffs","Inconsistent output quality",
         "Error-prone process","Customer-facing delay","Internal decision delay",
         "Compliance/privacy concern","Other"]
col_D = ["AI-assisted writing","AI-assisted summarization",
         "AI-assisted research/retrieval","Reporting automation",
         "Data pipeline/integration","Workflow automation",
         "Template/process redesign","Knowledge management improvement",
         "Human-in-the-loop review process","Not enough information","Not a fit"]
col_E = ["New","Needs Discovery","Discovery Scheduled","Discovery Complete",
         "Prototype Candidate","In Prototype","Implemented","Parked","Not a Fit"]
col_F = ["Low","Medium","High","Unknown"]
col_G = ["Easy","Medium","Hard","Unknown"]
col_H = ["Quick Win","Discovery Needed","Strategic Project",
         "High Risk / Needs Review","Not a Fit","Process Issue"]

lists_data = [col_A, col_B, col_C, col_D, col_E, col_F, col_G, col_H]
max_rows = max(len(c) for c in lists_data)
for r in range(max_rows):
    fill_hex = ROW_ALT if (r+2) % 2 == 1 else ROW_WHITE
    for col_idx, col_data in enumerate(lists_data, 1):
        cell = ws_lists.cell(r+2, col_idx)
        if r < len(col_data):
            cell.value = col_data[r]
        style_data(cell, fill_hex)

# Scoring weights table
style_header(ws_lists.cell(1, 10), "Setting")
style_header(ws_lists.cell(1, 11), "Value")

weights = [
    ("Impact Weight", 2),
    ("Frequency Weight", 1),
    ("Pain Weight", 2),
    ("Feasibility Weight", 1),
    ("Risk Penalty Weight", 2),
]
for i, (label, val) in enumerate(weights, 2):
    ws_lists.cell(i, 10).value = label
    ws_lists.cell(i, 10).font = Font(name='Calibri', size=10)
    ws_lists.cell(i, 10).fill = make_fill(ROW_ALT if i%2==1 else ROW_WHITE)
    ws_lists.cell(i, 10).border = make_border()
    ws_lists.cell(i, 11).value = val
    ws_lists.cell(i, 11).font = Font(name='Calibri', size=10)
    ws_lists.cell(i, 11).fill = make_fill(ROW_ALT if i%2==1 else ROW_WHITE)
    ws_lists.cell(i, 11).border = make_border()

ws_lists.merge_cells('J8:K8')
note_cell = ws_lists['J8']
note_cell.value = ("Note: Changing these weights will affect scoring behavior across the "
                   "Opportunity Scoring tab. Adjust with care.")
note_cell.font = Font(name='Calibri', size=10, italic=True)
note_cell.fill = make_fill(YELLOW_FILL)
note_cell.alignment = Alignment(wrap_text=True, vertical='top')
note_cell.border = make_border()

set_col_widths(ws_lists, {
    'A':22,'B':22,'C':32,'D':30,'E':24,'F':14,'G':14,'H':26,'J':24,'K':10
})

# ---------------------------------------------------------------------------
# SHEET 1: Intake Responses
# ---------------------------------------------------------------------------
ws_intake.sheet_properties.tabColor = "2E75B6"

intake_headers = [
    "Opportunity ID","Submitted Date","Submitted By","Team / Function",
    "Workflow / Task Name","Workflow Owner","People Involved","Who Uses the Output?",
    "Purpose of the Workflow","Decision / Deliverable / Outcome Supported",
    "What Happens If This Is Late Wrong or Skipped?","Current Process Summary",
    "Frequency","Minutes Per Run","People Doing This Work","Outputs / Items Per Cycle",
    "Systems / Tools Involved","Number of Systems","Friction Types","Pain Rating 1-5",
    "Business Criticality 1-5","Error-Proneness 1-5","Urgency 1-5","People Affected",
    "Customers / Districts / Students Affected","Desired Future State",
    "What Should Still Require Human Review?","What Should Never Be Automated?",
    "What Would Make You Trust the Improved Workflow?",
    "What Would Make This Unusable?","How Would We Know This Worked?",
    "Sensitive Data?","Data Types Involved","Input Data Location","Output Destination",
    "Existing Reports / Exports / Templates?","Example Link",
    "Available for Live Walkthrough?","Prototype Willingness","Ideal Timeline",
    "Initial Notes"
]
# 41 columns
for i, h in enumerate(intake_headers, 1):
    style_header(ws_intake.cell(1, i), h)

# Header notes
notes_intake = {
    'E': "Name the actual work being performed, not the proposed AI idea.",
    'I': "Explain why this workflow exists and what business outcome it supports.",
    'K': "Describe the consequence if this work is delayed, incorrect, incomplete, or skipped.",
    'L': "Describe what happens today. Do not solve here; capture current-state steps, handoffs, and manual effort.",
    'M': "How often this work occurs. Use Unknown if the submitter cannot estimate.",
    'N': "Approximate active minutes required each time the workflow runs.",
    'O': "Number of people who perform this workflow each cycle.",
    'S': "Types of pain present. Multiple values may apply.",
    'Z': "Describe what better would look like in terms of outcome and experience.",
    'AA': "Judgment, approvals, sensitivity checks, or quality review that should remain human-owned.",
    'AB': "Boundaries for automation or AI assistance.",
    'AC': "Controls, evidence, review steps, or transparency needed for adoption.",
    'AD': "Failure modes or missing requirements that would make the solution unacceptable.",
    'AE': "Measurable or observable success criteria.",
    'AF': "Whether sensitive, private, regulated, customer, student, employee, financial, legal, or confidential data may be involved.",
    'AG': "Specific kinds of data used by the workflow.",
    'AH': "Where source data lives today.",
    'AI': "Where the completed output goes.",
}
for col_letter, note_text in notes_intake.items():
    add_note(ws_intake[f'{col_letter}1'], note_text)

# Formulas rows 2–501
for row in range(2, 502):
    ws_intake.cell(row, 1).value = f'="WF-"&TEXT(ROW()-1,"0000")'

# Data validations
dv_team = make_dv("'Lists & Settings'!$A$2:$A$15")
dv_team.sqref = f"D2:D501"
ws_intake.add_data_validation(dv_team)

dv_freq = make_dv("'Lists & Settings'!$B$2:$B$8")
dv_freq.sqref = "M2:M501"
ws_intake.add_data_validation(dv_freq)

dv_sensitive = make_dv('"Yes,No,Unsure"')
dv_sensitive.sqref = "AF2:AF501"
ws_intake.add_data_validation(dv_sensitive)

dv_walkthrough = make_dv('"Yes,No,Maybe"')
dv_walkthrough.sqref = "AL2:AL501"
ws_intake.add_data_validation(dv_walkthrough)

dv_proto = make_dv('"Yes,No,Maybe"')
dv_proto.sqref = "AM2:AM501"
ws_intake.add_data_validation(dv_proto)

dv_timeline = make_dv('"This week,This month,This quarter,No specific timeline"')
dv_timeline.sqref = "AN2:AN501"
ws_intake.add_data_validation(dv_timeline)

# Alternating rows
apply_alt_rows(ws_intake, 2, 501, 41)

# Row heights
for r in range(2, 502):
    ws_intake.row_dimensions[r].height = 25

ws_intake.freeze_panes = 'A2'
ws_intake.auto_filter.ref = f"A1:{get_column_letter(41)}1"

intake_widths = {
    'A':14,'B':14,'C':20,'D':20,'E':30,'F':22,'G':20,'H':22,'I':35,'J':35,
    'K':35,'L':40,'M':18,'N':16,'O':18,'P':18,'Q':30,'R':16,'S':30,'T':14,
    'U':18,'V':18,'W':12,'X':16,'Y':24,'Z':35,'AA':35,'AB':35,'AC':35,
    'AD':35,'AE':35,'AF':14,'AG':25,'AH':25,'AI':25,'AJ':25,'AK':25,
    'AL':20,'AM':18,'AN':20,'AO':35
}
set_col_widths(ws_intake, intake_widths)

# ---------------------------------------------------------------------------
# SHEET 2: Opportunity Scoring
# ---------------------------------------------------------------------------
ws_scoring.sheet_properties.tabColor = "375623"

scoring_headers = [
    "Opportunity ID","Workflow / Task Name","Team / Function","Workflow Owner",
    "Status","Frequency","Minutes Per Run","People Doing This Work",
    "Estimated Runs Per Month","Estimated Monthly Hours","Annualized Hours",
    "Pain Rating","Business Criticality","Error-Proneness","Urgency",
    "People Affected","Sensitive Data?","Friction Types","Likely Solution Type",
    "Feasibility","Risk Level","Impact Score","Frequency Score","Friction Score",
    "Risk Penalty","Priority Score","Priority Category","Recommended Next Step",
    "Assignee","Next Review Date","Notes"
]
# 31 columns
for i, h in enumerate(scoring_headers, 1):
    style_header(ws_scoring.cell(1, i), h)

# Header notes for scoring
scoring_notes = {
    'E': "Current stage in the opportunity pipeline.",
    'S': "Best current guess at what type of improvement may fit. Update manually after triage.",
    'T': "Estimated implementation difficulty. Update manually after triage.",
    'U': "Risk based on data sensitivity, compliance, operational consequences, and customer impact.",
    'V': "Formula-derived score combining business criticality, urgency, and estimated monthly hours.",
    'W': "Formula-derived frequency score (how often the workflow occurs).",
    'X': "Average of pain rating and error-proneness.",
    'Y': "Penalty for high risk level and sensitive data involvement.",
    'Z': "Combined weighted priority score. Use to sort candidates, then apply judgment.",
    'AA': "Formula-derived triage bucket. Can be overridden by reviewer.",
    'AB': "Suggested next action based on Priority Category.",
}
for col_letter, note_text in scoring_notes.items():
    add_note(ws_scoring[f'{col_letter}1'], note_text)

# Formulas — columns B,C,D,F,G,H,I,J,K,L,M,N,O,P,Q,R,V,W,X,Y,Z,AA,AB
# Note: E(5), S(19), T(20), U(21), AC(29), AD(30), AE(31) are manual

for row in range(2, 502):
    r = row
    # B=2 Workflow/Task Name
    ws_scoring.cell(r, 2).value  = f"=IFERROR(XLOOKUP(A{r},'Intake Responses'!$A:$A,'Intake Responses'!$E:$E),\"\")"
    # C=3 Team/Function
    ws_scoring.cell(r, 3).value  = f"=IFERROR(XLOOKUP(A{r},'Intake Responses'!$A:$A,'Intake Responses'!$D:$D),\"\")"
    # D=4 Workflow Owner
    ws_scoring.cell(r, 4).value  = f"=IFERROR(XLOOKUP(A{r},'Intake Responses'!$A:$A,'Intake Responses'!$F:$F),\"\")"
    # F=6 Frequency
    ws_scoring.cell(r, 6).value  = f"=IFERROR(XLOOKUP(A{r},'Intake Responses'!$A:$A,'Intake Responses'!$M:$M),\"\")"
    # G=7 Minutes Per Run
    ws_scoring.cell(r, 7).value  = f"=IFERROR(XLOOKUP(A{r},'Intake Responses'!$A:$A,'Intake Responses'!$N:$N),\"\")"
    # H=8 People Doing This Work
    ws_scoring.cell(r, 8).value  = f"=IFERROR(XLOOKUP(A{r},'Intake Responses'!$A:$A,'Intake Responses'!$O:$O),\"\")"
    # I=9 Estimated Runs Per Month
    ws_scoring.cell(r, 9).value  = (f'=IFERROR(SWITCH(F{r},"Multiple times per day",60,'
                                    f'"Daily",20,"Weekly",4,"Monthly",1,"Quarterly",0.33,'
                                    f'"Ad hoc",0.5,"Unknown",0),0)')
    # J=10 Estimated Monthly Hours
    ws_scoring.cell(r, 10).value = f"=IFERROR((G{r}*H{r}*I{r})/60,0)"
    # K=11 Annualized Hours
    ws_scoring.cell(r, 11).value = f"=IFERROR(J{r}*12,0)"
    # L=12 Pain Rating
    ws_scoring.cell(r, 12).value = f"=IFERROR(XLOOKUP(A{r},'Intake Responses'!$A:$A,'Intake Responses'!$T:$T),\"\")"
    # M=13 Business Criticality
    ws_scoring.cell(r, 13).value = f"=IFERROR(XLOOKUP(A{r},'Intake Responses'!$A:$A,'Intake Responses'!$U:$U),\"\")"
    # N=14 Error-Proneness
    ws_scoring.cell(r, 14).value = f"=IFERROR(XLOOKUP(A{r},'Intake Responses'!$A:$A,'Intake Responses'!$V:$V),\"\")"
    # O=15 Urgency
    ws_scoring.cell(r, 15).value = f"=IFERROR(XLOOKUP(A{r},'Intake Responses'!$A:$A,'Intake Responses'!$W:$W),\"\")"
    # P=16 People Affected
    ws_scoring.cell(r, 16).value = f"=IFERROR(XLOOKUP(A{r},'Intake Responses'!$A:$A,'Intake Responses'!$X:$X),\"\")"
    # Q=17 Sensitive Data?
    ws_scoring.cell(r, 17).value = f"=IFERROR(XLOOKUP(A{r},'Intake Responses'!$A:$A,'Intake Responses'!$AF:$AF),\"\")"
    # R=18 Friction Types
    ws_scoring.cell(r, 18).value = f"=IFERROR(XLOOKUP(A{r},'Intake Responses'!$A:$A,'Intake Responses'!$S:$S),\"\")"
    # V=22 Impact Score
    ws_scoring.cell(r, 22).value = (f'=IFERROR(MIN(5,ROUND(((M{r}+O{r})/2)+'
                                    f'(IF(J{r}>=20,2,IF(J{r}>=10,1.5,IF(J{r}>=5,1,'
                                    f'IF(J{r}>=1,0.5,0))))),1)),0)')
    # W=23 Frequency Score
    ws_scoring.cell(r, 23).value = (f'=IFERROR(SWITCH(F{r},"Multiple times per day",5,'
                                    f'"Daily",4,"Weekly",3,"Monthly",2,"Quarterly",1,'
                                    f'"Ad hoc",1,"Unknown",0),0)')
    # X=24 Friction Score
    ws_scoring.cell(r, 24).value = f"=IFERROR(ROUND(AVERAGE(L{r},N{r}),1),0)"
    # Y=25 Risk Penalty
    ws_scoring.cell(r, 25).value = (f'=IFERROR(SWITCH(U{r},"High",3,"Medium",2,"Low",0,'
                                    f'"Unknown",1,1)+IF(Q{r}="Yes",2,IF(Q{r}="Unsure",1,0)),0)')
    # Z=26 Priority Score
    ws_scoring.cell(r, 26).value = (f'=IFERROR(ROUND((V{r}*2)+(W{r}*1)+(X{r}*2)+'
                                    f'(IF(T{r}="Easy",2,IF(T{r}="Medium",1,IF(T{r}="Hard",-1,0))))-(Y{r}*2),1),0)')
    # AA=27 Priority Category
    ws_scoring.cell(r, 27).value = (f'=IFERROR(IF(S{r}="Not a fit","Not a Fit",'
                                    f'IF(Y{r}>=5,"High Risk / Needs Review",'
                                    f'IF(AND(Z{r}>=18,T{r}="Easy"),"Quick Win",'
                                    f'IF(AND(Z{r}>=18,T{r}<>"Easy"),"Strategic Project",'
                                    f'IF(Z{r}>=10,"Discovery Needed","Process Issue"))))),"")')
    # AB=28 Recommended Next Step
    ws_scoring.cell(r, 28).value = (f'=IFERROR(SWITCH(AA{r},'
                                    f'"Quick Win","Schedule prototype scoping conversation",'
                                    f'"Discovery Needed","Schedule workflow walkthrough",'
                                    f'"Strategic Project","Identify stakeholders and dependencies",'
                                    f'"High Risk / Needs Review","Review with Security/Legal/Data owner before solutioning",'
                                    f'"Process Issue","Explore process/template/documentation improvement",'
                                    f'"Not a Fit","Do not pursue at this time",'
                                    f'"Review manually"),"")')

# Data validations for scoring
dv_s_status = make_dv("'Lists & Settings'!$E$2:$E$10")
dv_s_status.sqref = "E2:E501"
ws_scoring.add_data_validation(dv_s_status)

dv_s_sol = make_dv("'Lists & Settings'!$D$2:$D$12")
dv_s_sol.sqref = "S2:S501"
ws_scoring.add_data_validation(dv_s_sol)

dv_s_feas = make_dv("'Lists & Settings'!$G$2:$G$5")
dv_s_feas.sqref = "T2:T501"
ws_scoring.add_data_validation(dv_s_feas)

dv_s_risk = make_dv("'Lists & Settings'!$F$2:$F$5")
dv_s_risk.sqref = "U2:U501"
ws_scoring.add_data_validation(dv_s_risk)

dv_s_cat = make_dv("'Lists & Settings'!$H$2:$H$7")
dv_s_cat.sqref = "AA2:AA501"
ws_scoring.add_data_validation(dv_s_cat)

# Conditional formatting — Z (col 26) Priority Score
ws_scoring.conditional_formatting.add("Z2:Z501",
    FormulaRule(formula=['Z2>=20'], fill=make_fill(GREEN_FILL), font=Font(color=GREEN_FONT)))
ws_scoring.conditional_formatting.add("Z2:Z501",
    FormulaRule(formula=['AND(Z2>=10,Z2<20)'], fill=make_fill(YELLOW_FILL), font=Font(color=YELLOW_FONT)))
ws_scoring.conditional_formatting.add("Z2:Z501",
    FormulaRule(formula=['AND(Z2<10,Z2<>"")'], fill=make_fill(RED_FILL), font=Font(color=RED_FONT)))

# Conditional formatting — U (col 21) Risk Level
def make_text_cf_rule(col, text_val, fill_hex, font_hex):
    return FormulaRule(formula=[f'{col}2="{text_val}"'], fill=make_fill(fill_hex), font=Font(color=font_hex))

ws_scoring.conditional_formatting.add("U2:U501", make_text_cf_rule("U", "High", RED_FILL, RED_FONT))
ws_scoring.conditional_formatting.add("U2:U501", make_text_cf_rule("U", "Medium", YELLOW_FILL, YELLOW_FONT))
ws_scoring.conditional_formatting.add("U2:U501", make_text_cf_rule("U", "Low", GREEN_FILL, GREEN_FONT))
ws_scoring.conditional_formatting.add("U2:U501", make_text_cf_rule("U", "Unknown", GRAY_FILL, GRAY_FONT))

# Conditional formatting — AA (col 27) Priority Category
ws_scoring.conditional_formatting.add("AA2:AA501", make_text_cf_rule("AA", "Quick Win", GREEN_FILL, GREEN_FONT))
ws_scoring.conditional_formatting.add("AA2:AA501", make_text_cf_rule("AA", "Discovery Needed", BLUE_FILL, BLUE_FONT))
ws_scoring.conditional_formatting.add("AA2:AA501", make_text_cf_rule("AA", "Strategic Project", PURPLE_FILL, PURPLE_FONT))
ws_scoring.conditional_formatting.add("AA2:AA501", make_text_cf_rule("AA", "High Risk / Needs Review", RED_FILL, RED_FONT))
ws_scoring.conditional_formatting.add("AA2:AA501", make_text_cf_rule("AA", "Process Issue", YELLOW_FILL, YELLOW_FONT))
ws_scoring.conditional_formatting.add("AA2:AA501", make_text_cf_rule("AA", "Not a Fit", GRAY_FILL, GRAY_FONT))

# Row heights & alternating rows
apply_alt_rows(ws_scoring, 2, 501, 31)
for r in range(2, 502):
    ws_scoring.row_dimensions[r].height = 25

ws_scoring.freeze_panes = 'A2'
ws_scoring.auto_filter.ref = f"A1:{get_column_letter(31)}1"

scoring_widths = {
    'A':14,'B':28,'C':20,'D':22,'E':20,'F':18,'G':14,'H':16,'I':20,'J':20,
    'K':18,'L':14,'M':18,'N':16,'O':12,'P':16,'Q':14,'R':28,'S':28,'T':14,
    'U':14,'V':14,'W':16,'X':14,'Y':14,'Z':14,'AA':24,'AB':35,'AC':18,
    'AD':16,'AE':35
}
set_col_widths(ws_scoring, scoring_widths)

# ---------------------------------------------------------------------------
# SHEET 3: Discovery Sessions
# ---------------------------------------------------------------------------
ws_discovery.sheet_properties.tabColor = "7030A0"

discovery_headers = [
    "Discovery Session ID","Opportunity ID","Workflow / Task Name","Session Date",
    "Facilitator","Workflow Owner / SME","Session Type","Session Goal",
    "Summary of What We Learned","Key Pain Points","Risks / Guardrails",
    "Dependencies / Access Needed","Likely Solution Direction","Recommended Next Step",
    "Owner of Next Step","Next Review Date","Link to Detailed Step Log","Notes"
]
for i, h in enumerate(discovery_headers, 1):
    style_header(ws_discovery.cell(1, i), h)

# Header notes
disc_notes = {
    'A': "Auto-generated readable ID for this discovery conversation or session.",
    'B': "The Opportunity ID this discovery session is about.",
    'G': "The kind of discovery activity: Intake Review, Live Workflow Walkthrough, Stakeholder Interview, Async Review, Prototype Scoping, or Follow-up.",
    'H': "What this session is meant to clarify.",
    'I': "Brief narrative summary of the workflow, pain, constraints, or key discoveries.",
    'J': "Most important friction points found during discovery.",
    'K': "Privacy, compliance, security, data quality, customer impact, or human-review considerations.",
    'L': "Systems, permissions, APIs, reports, templates, stakeholders, or decisions needed before improvement work can proceed.",
    'N': "Concrete next action after this session.",
    'Q': "Optional link to filter the Discovery Step Log by this Opportunity ID.",
}
for col_letter, note_text in disc_notes.items():
    add_note(ws_discovery[f'{col_letter}1'], note_text)

# Formulas rows 2–501
for row in range(2, 502):
    r = row
    # A = Discovery Session ID
    ws_discovery.cell(r, 1).value = f'=IF(B{r}="","","DS-"&TEXT(ROW()-1,"0000"))'
    # C = Workflow / Task Name
    ws_discovery.cell(r, 3).value = f"=IFERROR(XLOOKUP(B{r},'Opportunity Scoring'!$A:$A,'Opportunity Scoring'!$B:$B),\"\")"
    # F = Workflow Owner / SME
    ws_discovery.cell(r, 6).value = f"=IFERROR(XLOOKUP(B{r},'Opportunity Scoring'!$A:$A,'Opportunity Scoring'!$D:$D),\"\")"

# Data validations
dv_d_type = make_dv('"Intake Review,Live Workflow Walkthrough,Stakeholder Interview,Async Review,Prototype Scoping,Follow-up"')
dv_d_type.sqref = "G2:G501"
ws_discovery.add_data_validation(dv_d_type)

dv_d_sol = make_dv("'Lists & Settings'!$D$2:$D$12")
dv_d_sol.sqref = "M2:M501"
ws_discovery.add_data_validation(dv_d_sol)

apply_alt_rows(ws_discovery, 2, 501, 18)
for r in range(2, 502):
    ws_discovery.row_dimensions[r].height = 25

ws_discovery.freeze_panes = 'A2'
ws_discovery.auto_filter.ref = f"A1:{get_column_letter(18)}1"

disc_widths = {
    'A':18,'B':14,'C':28,'D':14,'E':18,'F':22,'G':22,'H':35,'I':45,'J':40,
    'K':40,'L':40,'M':28,'N':35,'O':20,'P':16,'Q':30,'R':35
}
set_col_widths(ws_discovery, disc_widths)

# ---------------------------------------------------------------------------
# SHEET 4: Discovery Step Log
# ---------------------------------------------------------------------------
ws_steplog.sheet_properties.tabColor = "9933CC"

steplog_headers = [
    "Opportunity ID","Discovery Session ID","Discovery Date","Facilitator",
    "Workflow Owner","Current Workflow Step #","Current Workflow Step Description",
    "Tool / System Used","Input","Output","Manual Action Taken","Time Spent on Step",
    "Pain Point Observed","Error / Risk Point","Human Judgment Required?",
    "Could This Be Automated?","Could This Be AI-Assisted?",
    "Dependency / Access Needed","Notes"
]
for i, h in enumerate(steplog_headers, 1):
    style_header(ws_steplog.cell(1, i), h)

# Header notes
steplog_notes = {
    'A': "Detailed step-level log. Many rows per opportunity. Filter by Opportunity ID to see one workflow's steps.",
    'F': "Sequential step number in the current-state workflow.",
    'G': "What happens in this step. Capture the action, not a proposed solution.",
    'H': "Tool, system, document, report, inbox, or file used in this step.",
    'I': "What information, file, request, data, or context enters this step.",
    'J': "What this step produces or changes.",
    'K': "What the human actually does: copy, paste, search, rewrite, verify, reconcile, approve, escalate, etc.",
    'M': "Pain, delay, confusion, repetition, missing context, or frustration observed.",
    'N': "Where errors, privacy issues, customer impact, compliance concerns, or rework could occur.",
    'O': "Whether this step requires human interpretation, accountability, empathy, approval, or judgment.",
    'P': "Whether deterministic automation could handle this step.",
    'Q': "Whether AI could assist with drafting, summarizing, retrieval, classification, checking, or transformation.",
    'R': "What would be required to improve this step.",
}
for col_letter, note_text in steplog_notes.items():
    add_note(ws_steplog[f'{col_letter}1'], note_text)

# Data validations
dv_sl_human = make_dv('"Yes,No,Sometimes,Unknown"')
dv_sl_human.sqref = "O2:O1001"
ws_steplog.add_data_validation(dv_sl_human)

dv_sl_auto = make_dv('"Yes,No,Partially,Unknown"')
dv_sl_auto.sqref = "P2:P1001"
ws_steplog.add_data_validation(dv_sl_auto)

dv_sl_ai = make_dv('"Yes,No,Partially,Unknown"')
dv_sl_ai.sqref = "Q2:Q1001"
ws_steplog.add_data_validation(dv_sl_ai)

apply_alt_rows(ws_steplog, 2, 1001, 19)
for r in range(2, 1002):
    ws_steplog.row_dimensions[r].height = 25

ws_steplog.freeze_panes = 'A2'
ws_steplog.auto_filter.ref = f"A1:{get_column_letter(19)}1"

steplog_widths = {
    'A':14,'B':16,'C':14,'D':16,'E':18,'F':16,'G':40,'H':25,'I':30,'J':30,
    'K':30,'L':16,'M':35,'N':35,'O':20,'P':20,'Q':20,'R':30,'S':30
}
set_col_widths(ws_steplog, steplog_widths)

# ---------------------------------------------------------------------------
# SHEET 5: Opportunity Briefs
# ---------------------------------------------------------------------------
ws_briefs.sheet_properties.tabColor = "1F4E79"

# Row 1: title
ws_briefs.merge_cells('A1:F1')
style_header(ws_briefs['A1'], "Opportunity Brief Generator", size=14)

# Row 2: ID input
ws_briefs['A2'].value = "Selected Opportunity ID:"
ws_briefs['A2'].font = Font(name='Calibri', bold=True, size=10)
ws_briefs['A2'].fill = make_fill(ROW_ALT)
ws_briefs['A2'].alignment = Alignment(horizontal='right', vertical='center')
ws_briefs['A2'].border = make_border()

ws_briefs['B2'].fill = make_fill(YELLOW_FILL)
ws_briefs['B2'].border = make_border()
ws_briefs['B2'].alignment = Alignment(vertical='center')
add_note(ws_briefs['B2'], "Enter an Opportunity ID (e.g., WF-0001)")

ws_briefs.merge_cells('C2:F2')
ws_briefs['C2'].value = "Enter an Opportunity ID from Opportunity Scoring (e.g., WF-0001)"
ws_briefs['C2'].font = Font(name='Calibri', italic=True, color=GRAY_FONT, size=10)
ws_briefs['C2'].fill = make_fill(ROW_WHITE)
ws_briefs['C2'].alignment = Alignment(vertical='center')
ws_briefs['C2'].border = make_border()

# Row 3: description
ws_briefs.merge_cells('A3:F3')
ws_briefs['A3'].value = ("This tab generates a structured opportunity brief from the selected "
                         "Opportunity ID. Use it for stakeholder review, discovery planning, or prototype scoping.")
ws_briefs['A3'].font = Font(name='Calibri', italic=True, color=GRAY_FONT, size=10)
ws_briefs['A3'].fill = make_fill("F2F2F2")
ws_briefs['A3'].alignment = Alignment(wrap_text=True, vertical='center')
ws_briefs['A3'].border = make_border()
ws_briefs.row_dimensions[3].height = 35

def brief_section_header(ws, row, text):
    ws.merge_cells(f'A{row}:F{row}')
    style_header(ws[f'A{row}'], text)
    ws.row_dimensions[row].height = 20

def brief_row(ws, row, label, formula):
    ws[f'A{row}'].value = label
    ws[f'A{row}'].font = Font(name='Calibri', bold=True, size=10)
    ws[f'A{row}'].fill = make_fill(ROW_ALT)
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws[f'A{row}'].border = make_border()
    ws.merge_cells(f'B{row}:F{row}')
    ws[f'B{row}'].value = formula
    ws[f'B{row}'].font = Font(name='Calibri', size=10)
    ws[f'B{row}'].fill = make_fill(ROW_WHITE)
    ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws[f'B{row}'].border = make_border()
    ws.row_dimensions[row].height = 40

# SECTION: OPPORTUNITY OVERVIEW (row 5)
brief_section_header(ws_briefs, 5, "OPPORTUNITY OVERVIEW")
brief_row(ws_briefs, 6, "Opportunity ID", "=$B$2")
brief_row(ws_briefs, 7, "Workflow / Task Name", "=IFERROR(XLOOKUP($B$2,'Opportunity Scoring'!$A:$A,'Opportunity Scoring'!$B:$B),\"\")")
brief_row(ws_briefs, 8, "Team / Function", "=IFERROR(XLOOKUP($B$2,'Opportunity Scoring'!$A:$A,'Opportunity Scoring'!$C:$C),\"\")")
brief_row(ws_briefs, 9, "Workflow Owner", "=IFERROR(XLOOKUP($B$2,'Opportunity Scoring'!$A:$A,'Opportunity Scoring'!$D:$D),\"\")")
brief_row(ws_briefs, 10, "Status", "=IFERROR(XLOOKUP($B$2,'Opportunity Scoring'!$A:$A,'Opportunity Scoring'!$E:$E),\"\")")
brief_row(ws_briefs, 11, "Purpose of the Workflow", "=IFERROR(XLOOKUP($B$2,'Intake Responses'!$A:$A,'Intake Responses'!$I:$I),\"\")")

# blank separator
ws_briefs.row_dimensions[12].height = 8

# SECTION: FRICTION & TIME COST (row 13)
brief_section_header(ws_briefs, 13, "FRICTION & TIME COST")
brief_row(ws_briefs, 14, "Current Friction Summary", "=IFERROR(XLOOKUP($B$2,'Intake Responses'!$A:$A,'Intake Responses'!$L:$L),\"\")")
brief_row(ws_briefs, 15, "Friction Types", "=IFERROR(XLOOKUP($B$2,'Intake Responses'!$A:$A,'Intake Responses'!$S:$S),\"\")")
brief_row(ws_briefs, 16, "Pain Rating (1-5)", "=IFERROR(XLOOKUP($B$2,'Opportunity Scoring'!$A:$A,'Opportunity Scoring'!$L:$L),\"\")")
brief_row(ws_briefs, 17, "Frequency", "=IFERROR(XLOOKUP($B$2,'Opportunity Scoring'!$A:$A,'Opportunity Scoring'!$F:$F),\"\")")
brief_row(ws_briefs, 18, "Minutes Per Run", "=IFERROR(XLOOKUP($B$2,'Opportunity Scoring'!$A:$A,'Opportunity Scoring'!$G:$G),\"\")")
brief_row(ws_briefs, 19, "People Doing This Work", "=IFERROR(XLOOKUP($B$2,'Opportunity Scoring'!$A:$A,'Opportunity Scoring'!$H:$H),\"\")")
brief_row(ws_briefs, 20, "Est. Monthly Hours", "=IFERROR(XLOOKUP($B$2,'Opportunity Scoring'!$A:$A,'Opportunity Scoring'!$J:$J),\"\")")
brief_row(ws_briefs, 21, "Annualized Hours", "=IFERROR(XLOOKUP($B$2,'Opportunity Scoring'!$A:$A,'Opportunity Scoring'!$K:$K),\"\")")

ws_briefs.row_dimensions[22].height = 8

# SECTION: BUSINESS IMPACT (row 23)
brief_section_header(ws_briefs, 23, "BUSINESS IMPACT")
brief_row(ws_briefs, 24, "Business Criticality (1-5)", "=IFERROR(XLOOKUP($B$2,'Opportunity Scoring'!$A:$A,'Opportunity Scoring'!$M:$M),\"\")")
brief_row(ws_briefs, 25, "Urgency (1-5)", "=IFERROR(XLOOKUP($B$2,'Opportunity Scoring'!$A:$A,'Opportunity Scoring'!$O:$O),\"\")")
brief_row(ws_briefs, 26, "What Happens If Late/Wrong/Skipped", "=IFERROR(XLOOKUP($B$2,'Intake Responses'!$A:$A,'Intake Responses'!$K:$K),\"\")")
brief_row(ws_briefs, 27, "Internal People Affected", "=IFERROR(XLOOKUP($B$2,'Opportunity Scoring'!$A:$A,'Opportunity Scoring'!$P:$P),\"\")")
brief_row(ws_briefs, 28, "Customers / Districts / Students Affected", "=IFERROR(XLOOKUP($B$2,'Intake Responses'!$A:$A,'Intake Responses'!$Y:$Y),\"\")")

ws_briefs.row_dimensions[29].height = 8

# SECTION: RISK & DATA (row 30)
brief_section_header(ws_briefs, 30, "RISK & DATA CONSIDERATIONS")
brief_row(ws_briefs, 31, "Sensitive Data?", "=IFERROR(XLOOKUP($B$2,'Opportunity Scoring'!$A:$A,'Opportunity Scoring'!$Q:$Q),\"\")")
brief_row(ws_briefs, 32, "Risk Level", "=IFERROR(XLOOKUP($B$2,'Opportunity Scoring'!$A:$A,'Opportunity Scoring'!$U:$U),\"\")")
brief_row(ws_briefs, 33, "Data Types Involved", "=IFERROR(XLOOKUP($B$2,'Intake Responses'!$A:$A,'Intake Responses'!$AG:$AG),\"\")")
brief_row(ws_briefs, 34, "Input Data Location", "=IFERROR(XLOOKUP($B$2,'Intake Responses'!$A:$A,'Intake Responses'!$AH:$AH),\"\")")
brief_row(ws_briefs, 35, "Output Destination", "=IFERROR(XLOOKUP($B$2,'Intake Responses'!$A:$A,'Intake Responses'!$AI:$AI),\"\")")

ws_briefs.row_dimensions[36].height = 8

# SECTION: LIKELY SOLUTION DIRECTION (row 37)
brief_section_header(ws_briefs, 37, "LIKELY SOLUTION DIRECTION")
brief_row(ws_briefs, 38, "Likely Solution Type", "=IFERROR(XLOOKUP($B$2,'Opportunity Scoring'!$A:$A,'Opportunity Scoring'!$S:$S),\"\")")
brief_row(ws_briefs, 39, "Feasibility", "=IFERROR(XLOOKUP($B$2,'Opportunity Scoring'!$A:$A,'Opportunity Scoring'!$T:$T),\"\")")
brief_row(ws_briefs, 40, "Recommended Next Step", "=IFERROR(XLOOKUP($B$2,'Opportunity Scoring'!$A:$A,'Opportunity Scoring'!$AB:$AB),\"\")")

ws_briefs.row_dimensions[41].height = 8

# SECTION: HUMAN REVIEW & GUARDRAILS (row 42)
brief_section_header(ws_briefs, 42, "HUMAN REVIEW & GUARDRAILS")
brief_row(ws_briefs, 43, "Should Still Require Human Review", "=IFERROR(XLOOKUP($B$2,'Intake Responses'!$A:$A,'Intake Responses'!$AA:$AA),\"\")")
brief_row(ws_briefs, 44, "Should Never Be Automated", "=IFERROR(XLOOKUP($B$2,'Intake Responses'!$A:$A,'Intake Responses'!$AB:$AB),\"\")")
brief_row(ws_briefs, 45, "What Would Make You Trust This", "=IFERROR(XLOOKUP($B$2,'Intake Responses'!$A:$A,'Intake Responses'!$AC:$AC),\"\")")
brief_row(ws_briefs, 46, "What Would Make This Unusable", "=IFERROR(XLOOKUP($B$2,'Intake Responses'!$A:$A,'Intake Responses'!$AD:$AD),\"\")")

ws_briefs.row_dimensions[47].height = 8

# SECTION: SUCCESS MEASUREMENT (row 48)
brief_section_header(ws_briefs, 48, "SUCCESS MEASUREMENT")
brief_row(ws_briefs, 49, "How Would We Know This Worked", "=IFERROR(XLOOKUP($B$2,'Intake Responses'!$A:$A,'Intake Responses'!$AE:$AE),\"\")")
brief_row(ws_briefs, 50, "Desired Future State", "=IFERROR(XLOOKUP($B$2,'Intake Responses'!$A:$A,'Intake Responses'!$Z:$Z),\"\")")

ws_briefs.row_dimensions[51].height = 8

# SECTION: PRIORITY ASSESSMENT (row 52)
brief_section_header(ws_briefs, 52, "PRIORITY ASSESSMENT")
brief_row(ws_briefs, 53, "Priority Score", "=IFERROR(XLOOKUP($B$2,'Opportunity Scoring'!$A:$A,'Opportunity Scoring'!$Z:$Z),\"\")")
brief_row(ws_briefs, 54, "Priority Category", "=IFERROR(XLOOKUP($B$2,'Opportunity Scoring'!$A:$A,'Opportunity Scoring'!$AA:$AA),\"\")")
brief_row(ws_briefs, 55, "Status", "=IFERROR(XLOOKUP($B$2,'Opportunity Scoring'!$A:$A,'Opportunity Scoring'!$E:$E),\"\")")

brief_widths = {'A':32,'B':16,'C':16,'D':16,'E':16,'F':16}
set_col_widths(ws_briefs, brief_widths)

# ---------------------------------------------------------------------------
# SHEET 6: Dashboard
# ---------------------------------------------------------------------------
ws_dashboard.sheet_properties.tabColor = "C55A11"

# Row 1 title
ws_dashboard.merge_cells('A1:L1')
style_header(ws_dashboard['A1'], "AI Enablement Opportunity Pipeline — Dashboard",
             size=14, bold=True)
ws_dashboard.row_dimensions[1].height = 30

# KPI section
def kpi_label(ws, row, text):
    ws.merge_cells(f'A{row}:C{row}')
    c = ws[f'A{row}']
    c.value = text
    c.font = Font(name='Calibri', bold=True, color=NAVY_FONT, size=11)
    c.fill = make_fill(NAVY)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    c.border = make_border()
    ws.row_dimensions[row].height = 22

def kpi_value(ws, row, formula):
    ws.merge_cells(f'D{row}:F{row}')
    c = ws[f'D{row}']
    c.value = formula
    c.font = Font(name='Calibri', bold=True, color=NAVY, size=13)
    c.fill = make_fill(ROW_ALT)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = make_border()

kpis = [
    (3, "Total Opportunities",           "=COUNTA('Opportunity Scoring'!$A$2:$A$501)"),
    (4, "Quick Wins",                    '=COUNTIF(\'Opportunity Scoring\'!$AA$2:$AA$501,"Quick Win")'),
    (5, "Discovery Needed",              '=COUNTIF(\'Opportunity Scoring\'!$AA$2:$AA$501,"Discovery Needed")'),
    (6, "Strategic Projects",            '=COUNTIF(\'Opportunity Scoring\'!$AA$2:$AA$501,"Strategic Project")'),
    (7, "High Risk / Needs Review",      '=COUNTIF(\'Opportunity Scoring\'!$AA$2:$AA$501,"High Risk / Needs Review")'),
    (9, "Est. Monthly Hours (All)",      "=ROUND(SUM('Opportunity Scoring'!$J$2:$J$501),1)"),
    (10,"Est. Annual Hours (All)",       "=ROUND(SUM('Opportunity Scoring'!$K$2:$K$501),1)"),
    (11,"Avg Priority Score",            '=IFERROR(ROUND(AVERAGEIF(\'Opportunity Scoring\'!$Z$2:$Z$501,">0"),1),"—")'),
    (12,"Discovery Sessions Logged",     "=COUNTA('Discovery Sessions'!$A$2:$A$501)"),
]
for row, label, formula in kpis:
    kpi_label(ws_dashboard, row, label)
    kpi_value(ws_dashboard, row, formula)

# Row 14: Portfolio Summary Tables header
ws_dashboard.merge_cells('A14:L14')
style_header(ws_dashboard['A14'], "Portfolio Summary Tables", size=11, bold=True)
ws_dashboard.row_dimensions[14].height = 22

# Helper: section sub-header
def dash_sub_header(ws, cell_ref, text, merge_to=None):
    if merge_to:
        ws.merge_cells(f'{cell_ref}:{merge_to}')
    c = ws[cell_ref]
    c.value = text
    c.font = Font(name='Calibri', bold=True, color=NAVY_FONT, size=10)
    c.fill = make_fill(MID_BLUE)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = make_border()

def dash_table_header(ws, row, col, text):
    c = ws.cell(row, col)
    c.value = text
    c.font = Font(name='Calibri', bold=True, size=10)
    c.fill = make_fill(ROW_ALT)
    c.alignment = Alignment(wrap_text=True, vertical='center')
    c.border = make_border()

def dash_table_data(ws, row, col, value, fill_hex=ROW_WHITE):
    c = ws.cell(row, col)
    c.value = value
    c.font = Font(name='Calibri', size=10)
    c.fill = make_fill(fill_hex)
    c.alignment = Alignment(vertical='center')
    c.border = make_border()

# Priority Category table (rows 16–23, col A)
dash_sub_header(ws_dashboard, 'A16', "Opportunities by Priority Category", 'C16')
ws_dashboard.row_dimensions[16].height = 20
dash_table_header(ws_dashboard, 17, 1, "Priority Category")
dash_table_header(ws_dashboard, 17, 2, "Count")
pcat_rows = [
    ("Quick Win",               '=COUNTIF(\'Opportunity Scoring\'!$AA$2:$AA$501,"Quick Win")'),
    ("Discovery Needed",        '=COUNTIF(\'Opportunity Scoring\'!$AA$2:$AA$501,"Discovery Needed")'),
    ("Strategic Project",       '=COUNTIF(\'Opportunity Scoring\'!$AA$2:$AA$501,"Strategic Project")'),
    ("High Risk / Needs Review",'=COUNTIF(\'Opportunity Scoring\'!$AA$2:$AA$501,"High Risk / Needs Review")'),
    ("Process Issue",           '=COUNTIF(\'Opportunity Scoring\'!$AA$2:$AA$501,"Process Issue")'),
    ("Not a Fit",               '=COUNTIF(\'Opportunity Scoring\'!$AA$2:$AA$501,"Not a Fit")'),
]
for i, (label, formula) in enumerate(pcat_rows, 18):
    fill = ROW_ALT if i % 2 == 1 else ROW_WHITE
    dash_table_data(ws_dashboard, i, 1, label, fill)
    dash_table_data(ws_dashboard, i, 2, formula, fill)

# Risk Level table (rows 16–21, col D)
dash_sub_header(ws_dashboard, 'D16', "Opportunities by Risk Level", 'F16')
dash_table_header(ws_dashboard, 17, 4, "Risk Level")
dash_table_header(ws_dashboard, 17, 5, "Count")
risk_rows = [
    ("High",    '=COUNTIF(\'Opportunity Scoring\'!$U$2:$U$501,"High")'),
    ("Medium",  '=COUNTIF(\'Opportunity Scoring\'!$U$2:$U$501,"Medium")'),
    ("Low",     '=COUNTIF(\'Opportunity Scoring\'!$U$2:$U$501,"Low")'),
    ("Unknown", '=COUNTIF(\'Opportunity Scoring\'!$U$2:$U$501,"Unknown")'),
]
for i, (label, formula) in enumerate(risk_rows, 18):
    fill = ROW_ALT if i % 2 == 1 else ROW_WHITE
    dash_table_data(ws_dashboard, i, 4, label, fill)
    dash_table_data(ws_dashboard, i, 5, formula, fill)

# Discovery Sessions table (rows 25–32, col A)
dash_sub_header(ws_dashboard, 'A25', "Discovery Sessions by Type", 'C25')
dash_table_header(ws_dashboard, 26, 1, "Session Type")
dash_table_header(ws_dashboard, 26, 2, "Count")
session_rows = [
    ("Intake Review",             '=COUNTIF(\'Discovery Sessions\'!$G$2:$G$501,"Intake Review")'),
    ("Live Workflow Walkthrough",  '=COUNTIF(\'Discovery Sessions\'!$G$2:$G$501,"Live Workflow Walkthrough")'),
    ("Stakeholder Interview",      '=COUNTIF(\'Discovery Sessions\'!$G$2:$G$501,"Stakeholder Interview")'),
    ("Async Review",               '=COUNTIF(\'Discovery Sessions\'!$G$2:$G$501,"Async Review")'),
    ("Prototype Scoping",          '=COUNTIF(\'Discovery Sessions\'!$G$2:$G$501,"Prototype Scoping")'),
    ("Follow-up",                  '=COUNTIF(\'Discovery Sessions\'!$G$2:$G$501,"Follow-up")'),
]
for i, (label, formula) in enumerate(session_rows, 27):
    fill = ROW_ALT if i % 2 == 1 else ROW_WHITE
    dash_table_data(ws_dashboard, i, 1, label, fill)
    dash_table_data(ws_dashboard, i, 2, formula, fill)

# Solution Type table (rows 25–35, col D)
dash_sub_header(ws_dashboard, 'D25', "Opportunities by Solution Type", 'G25')
dash_table_header(ws_dashboard, 26, 4, "Solution Type")
dash_table_header(ws_dashboard, 26, 5, "Count")
sol_rows = [
    ("AI-assisted writing",                  '=COUNTIF(\'Opportunity Scoring\'!$S$2:$S$501,"AI-assisted writing")'),
    ("AI-assisted summarization",            '=COUNTIF(\'Opportunity Scoring\'!$S$2:$S$501,"AI-assisted summarization")'),
    ("AI-assisted research/retrieval",       '=COUNTIF(\'Opportunity Scoring\'!$S$2:$S$501,"AI-assisted research/retrieval")'),
    ("Reporting automation",                 '=COUNTIF(\'Opportunity Scoring\'!$S$2:$S$501,"Reporting automation")'),
    ("Data pipeline/integration",            '=COUNTIF(\'Opportunity Scoring\'!$S$2:$S$501,"Data pipeline/integration")'),
    ("Workflow automation",                  '=COUNTIF(\'Opportunity Scoring\'!$S$2:$S$501,"Workflow automation")'),
    ("Template/process redesign",            '=COUNTIF(\'Opportunity Scoring\'!$S$2:$S$501,"Template/process redesign")'),
    ("Knowledge management improvement",     '=COUNTIF(\'Opportunity Scoring\'!$S$2:$S$501,"Knowledge management improvement")'),
    ("Human-in-the-loop review process",     '=COUNTIF(\'Opportunity Scoring\'!$S$2:$S$501,"Human-in-the-loop review process")'),
]
for i, (label, formula) in enumerate(sol_rows, 27):
    fill = ROW_ALT if i % 2 == 1 else ROW_WHITE
    dash_table_data(ws_dashboard, i, 4, label, fill)
    dash_table_data(ws_dashboard, i, 5, formula, fill)

# Row 38: note
ws_dashboard.merge_cells('A38:L38')
c = ws_dashboard['A38']
c.value = ("This is a portfolio/program-level view. For details on any opportunity, "
           "see Opportunity Briefs or Opportunity Scoring.")
c.font = Font(name='Calibri', italic=True, color=GRAY_FONT, size=10)
c.fill = make_fill("F2F2F2")
c.alignment = Alignment(wrap_text=True, vertical='center')
c.border = make_border()
ws_dashboard.row_dimensions[38].height = 30

dash_widths = {'A':28,'B':12,'C':12,'D':32,'E':12,'F':12,'G':12}
set_col_widths(ws_dashboard, dash_widths)

# ---------------------------------------------------------------------------
# SHEET 8: Instructions
# ---------------------------------------------------------------------------
ws_instr.sheet_properties.tabColor = "595959"

# Row 1: title
ws_instr.merge_cells('A1:H1')
style_header(ws_instr['A1'],
             "Instructions — AI Enablement Workflow Friction Intake System",
             size=14, bold=True)
ws_instr.row_dimensions[1].height = 30

def instr_section(ws, row, text):
    ws.merge_cells(f'A{row}:H{row}')
    c = ws[f'A{row}']
    c.value = text
    c.font = Font(name='Calibri', bold=True, color=NAVY_FONT, size=10)
    c.fill = make_fill(MID_BLUE)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = make_border()
    ws.row_dimensions[row].height = 22

def instr_content(ws, row, text, height=40, col_span='A:H'):
    ws.merge_cells(f'A{row}:H{row}')
    c = ws[f'A{row}']
    c.value = text
    c.font = Font(name='Calibri', size=10)
    c.fill = make_fill(ROW_WHITE)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    c.border = make_border()
    ws.row_dimensions[row].height = height

def instr_table_header_row(ws, row, cols_texts):
    for col_idx, text in enumerate(cols_texts, 1):
        c = ws.cell(row, col_idx)
        c.value = text
        c.font = Font(name='Calibri', bold=True, size=10)
        c.fill = make_fill(NAVY)
        c.font = Font(name='Calibri', bold=True, color=NAVY_FONT, size=10)
        c.alignment = Alignment(wrap_text=True, vertical='center')
        c.border = make_border()
    ws.row_dimensions[row].height = 22

def instr_table_row(ws, row, values, height=40):
    fill = ROW_ALT if row % 2 == 1 else ROW_WHITE
    for col_idx, val in enumerate(values, 1):
        c = ws.cell(row, col_idx)
        c.value = val
        c.font = Font(name='Calibri', size=10)
        c.fill = make_fill(fill)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        c.border = make_border()
    ws.row_dimensions[row].height = height

# PURPOSE
instr_section(ws_instr, 3, "PURPOSE")
instr_content(ws_instr, 4, ("This workbook is used to identify, quantify, and prioritize workflow friction "
    "opportunities that may be reduced through AI, automation, process redesign, documentation improvements, "
    "reporting improvements, integrations, or other enablement work."), height=50)

# IMPORTANT PRINCIPLE
instr_section(ws_instr, 6, "IMPORTANT PRINCIPLE")
instr_content(ws_instr, 7, ("Do not ask only: 'What should AI automate?' — Ask: 'Where is work creating "
    "unnecessary friction, and what would change if that friction were reduced?' The purpose of this workbook "
    "is to understand workflow friction first. AI may or may not be the right solution."), height=55)

# HOW TO USE THIS WORKBOOK
instr_section(ws_instr, 9, "HOW TO USE THIS WORKBOOK")
steps = [
    "1. Enter or collect new opportunities in Intake Responses.",
    "2. Review and score each opportunity in Opportunity Scoring.",
    "3. For promising opportunities, log each discovery conversation in Discovery Sessions.",
    "4. During live walkthroughs, use Discovery Step Log to capture the current workflow step by step.",
    "5. Use Opportunity Briefs to generate a structured summary for stakeholders.",
    "6. Use Dashboard to review the overall opportunity pipeline.",
]
for i, step in enumerate(steps, 10):
    instr_content(ws_instr, i, step, height=30)

# SHEET GUIDE
instr_section(ws_instr, 17, "SHEET GUIDE")
instr_table_header_row(ws_instr, 18, ["Sheet / Area", "What it collects", "How to use it"])
sheet_guide = [
    ("Intake Responses",
     "Raw descriptions of workflow friction, impact, data context, desired future state, and willingness to participate.",
     "Use as the source of truth for what the submitter knows. Do not over-score if the workflow is unclear."),
    ("Opportunity Scoring",
     "Triage fields and formulas translating intake into estimated hours, impact, friction, risk, priority, and next step.",
     "Review formula outputs, then manually complete Status, Solution Type, Feasibility, Risk Level, Assignee, and notes."),
    ("Discovery Sessions",
     "One row per discovery conversation, walkthrough, interview, async review, or prototype scoping session.",
     "Use as the human-readable index of discovery work. Summarizes learning, next steps, owners, and dependencies."),
    ("Discovery Step Log",
     "Detailed step-by-step observations from live workflow walkthroughs. Many rows per opportunity.",
     "Filter by Opportunity ID to view one workflow. One row per step so pain, risks, and automation opportunities are specific."),
    ("Opportunity Briefs",
     "Stakeholder-friendly one-page summary generated from a selected Opportunity ID.",
     "Choose the ID in B2, use the brief for discovery, prototype scoping, or stakeholder review."),
    ("Dashboard",
     "Leadership/program overview of the opportunity pipeline.",
     "Use for portfolio review: volume, quick wins, risk items, time estimates, discovery activity, and solution mix."),
    ("Lists & Settings",
     "Reusable dropdown values and scoring weights.",
     "Only edit intentionally; changes affect validations and scoring behavior."),
]
for i, row_vals in enumerate(sheet_guide, 19):
    instr_table_row(ws_instr, i, row_vals, height=50)

# TRIAGE CATEGORIES
instr_section(ws_instr, 27, "TRIAGE CATEGORIES")
triage_rows = [
    "Quick Win — High impact, low complexity, low risk.",
    "Discovery Needed — Promising but not yet understood deeply enough.",
    "Strategic Project — High impact but requires significant coordination or integration.",
    "High Risk / Needs Review — Involves sensitive data, compliance, legal, security, or customer-facing risk.",
    "Process Issue — Better solved by documentation, process clarification, templates, or training.",
    "Not a Fit — Low impact, inappropriate risk, or not worth pursuing.",
]
for i, text in enumerate(triage_rows, 28):
    instr_content(ws_instr, i, text, height=30)

# SCORING INTERPRETATION
instr_section(ws_instr, 35, "SCORING INTERPRETATION")
instr_table_header_row(ws_instr, 36, ["Field", "Meaning", "How to interpret it"])
scoring_guide = [
    ("Impact Score", "Combines criticality, urgency, and monthly hours.", "Directional, not absolute."),
    ("Frequency Score", "Based on how often the workflow occurs.", "High-frequency annoyances can be strong opportunities even if each instance is small."),
    ("Friction Score", "Average of pain rating and error-proneness.", "High friction points to standardization, QA, or automation opportunities."),
    ("Risk Penalty", "Penalty for risk level and sensitive data.", "High risk means review before solutioning, not necessarily \"no.\""),
    ("Priority Score", "Combined weighted score after impact, frequency, friction, feasibility, and risk.", "Use to sort, then apply judgment."),
    ("Priority Category", "Formula-derived triage bucket.", "Use as a starting point; a reviewer can override."),
]
for i, row_vals in enumerate(scoring_guide, 37):
    instr_table_row(ws_instr, i, row_vals, height=40)

# AI ENABLEMENT GUIDANCE
instr_section(ws_instr, 44, "AI ENABLEMENT GUIDANCE")
ai_guidance = [
    ("Consider AI when the workflow involves: drafting, rewriting, summarizing, synthesizing, searching/retrieval, "
     "classification, pattern recognition, triage, repetitive judgment support, or transformation of messy inputs into structured outputs."),
    "",
    ("Treat AI carefully when the workflow involves: sensitive data, customer-facing decisions, legal, compliance, security, "
     "financial, HR, or student data, irreversible actions, high-stakes decisions, or areas where hallucination or missing "
     "context would cause harm."),
    "",
    ("Also consider: process redesign, templates, better documentation, reporting automation, system integrations, "
     "workflow automation, knowledge management improvements, human-in-the-loop review, or better ownership and handoff design."),
]
for i, text in enumerate(ai_guidance, 45):
    instr_content(ws_instr, i, text, height=50 if text else 10)

set_col_widths(ws_instr, {'A':28,'B':40,'C':40})

# ---------------------------------------------------------------------------
# Save
# ---------------------------------------------------------------------------
output_path = "/Users/tom.leger/repo/discovery-intake/AI_Enablement_Intake.xlsx"
wb.save(output_path)
size = os.path.getsize(output_path)
print(f"Saved: {output_path}")
print(f"File size: {size:,} bytes ({size/1024:.1f} KB)")
